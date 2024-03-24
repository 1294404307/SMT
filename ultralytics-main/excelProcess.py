import os
import openpyxl
path = r"F:\数模\C"
os.chdir(path)  # 修改工作路径

workbook2 = openpyxl.load_workbook('花菜类.xlsx')	# 返回一个workbook数据类型的值
sheet = workbook2['Sheet1']
row=1
sum={}
m=0
allday=[]
sumdanjia=0
z=1
danjia=[]
for i,j,k,p in zip(sheet['A'],sheet['D'],sheet['M'],sheet['E']):
    try:
        p.value+5
    except:
        continue
    if m==i.value:
        try:
            sum[k.value]+=j.value
        except:
            sum[k.value]=j.value
        sumdanjia+=float(p.value)
        z+=1
    else:
        allday.append(sum)
        danjia.append(sumdanjia/z)
        sum={}
        sum[k.value]=j.value
        m=i.value
        sumdanjia=p.value
        z=1
    print(i.value)
print(allday)

data = openpyxl.Workbook() # 新建工作簿
data.create_sheet('Sheet1') # 添加页
#table = data.get_sheet_by_name('Sheet1') # 获得指定名称页
table = data.active # 获得当前活跃的工作页，默认为第一个工作页
row=1
for i,s in zip(allday,danjia):
    for m in i:
        table.cell(row,1,m)
        table.cell(row, 2, i[m])
        table.cell(row, 3, s)
        row+=1
data.save('花菜类处理.xlsx')



# path = r"F:\数模\C"
# os.chdir(path)  # 修改工作路径
# #
# workbook1 = openpyxl.load_workbook('附件1.xlsx')	# 返回一个workbook数据类型的值
# # print(workbook.sheetnames)	# 打印Excel表中的所有表
# sheet1 = workbook1['Sheet1']
# # print(sheet1['A1'].value)
#
# workbook2 = openpyxl.load_workbook('附件4.xlsx')	# 返回一个workbook数据类型的值
# sheet2 = workbook2['Sheet1']
#
# codetoclass={}
# for i,j in zip(sheet1['A'],sheet1['D']):
#     codetoclass[i.value]=j.value
# sheet2['D1'].value='分类名称'
# k=1
# for i in sheet2['A']:
#     sheet2['D'+str(k)].value=codetoclass[i.value]
#     k+=1
# workbook2.save('pp.xlsx')
# workbook2.close()



# path = r"F:\数模\CCC"
# os.chdir(path)  # 修改工作路径
# #
# workbook1 = openpyxl.load_workbook('附件2.xlsx')	# 返回一个workbook数据类型的值
# # print(workbook.sheetnames)	# 打印Excel表中的所有表
# sheet1 = workbook1['Sheet2']
# sheet2 = workbook1['Sheet5']
#
# codetoclass={}
# for i,j,k in zip(sheet1['A'],sheet1['B'],sheet1['C']):
#     codetoclass[str(i.value)+j.value]=k.value
# sheet2['J1'].value='批发价格'
# k=1
# for i,j in zip(sheet2['A'],sheet2['C']):
#     try:
#         sheet2['J'+str(k)].value=codetoclass[str(i.value)+j.value]
#     except:
#         pass
#     k+=1
# workbook1.save('pp.xlsx')
# workbook1.close()

# path = r"F:\数模\CCC"
# os.chdir(path)  # 修改工作路径
# name='附件2'
# workbook1 = openpyxl.load_workbook(name+'.xlsx')	# 返回一个workbook数据类型的值
# # print(workbook.sheetnames)	# 打印Excel表中的所有表
# sheet1 = workbook1['Sheet2']
# sheet2 = workbook1['Sheet1']
#
# codetoclass={}
# for i,j in zip(sheet1['H'],sheet1['C']):
#     codetoclass[(i.value)]=j.value
# sheet2['K1'].value='商品编号'
# k=1
# for i,j in zip(sheet2['A'],sheet2['K']):
#     try:
#         sheet2['k'+str(k)].value=codetoclass[(i.value)]
#     except:
#         pass
#     k+=1
# workbook1.save('附件一处理.xlsx')
# workbook1.close()

# print(workbook2.sheetnames)	# 打印Excel表中的所有表
# sheet2 = workbook2.active
# # print(sheet2['A1'].value)
#
# sheet2['A']
# matoname={}
# i=sheet1['C']
# j=sheet1['D']
# for m,n in zip(i,j):
#     # matoname.append({m.value:n.value})
#     matoname[m.value]=n.value
#
# print(matoname)
#
# m=1
# for i in sheet2['C']:
#     sheet2['H'+str(m)]=matoname[i.value]
#     m+=1
#
# for i in sheet2['H']:
#     print(i.value)
# workbook2.save('s.xlsx')
# print(x.value)
# for i in sheet1[]





# 结果：
# ['Sheet1', 'Sheet2']
