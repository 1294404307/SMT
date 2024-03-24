import os
import openpyxl

path = r"F:\数模\C"
os.chdir(path)  # 修改工作路径
#
# workbook = openpyxl.load_workbook('附件1.xlsx')	# 返回一个workbook数据类型的值
# # print(workbook.sheetnames)	# 打印Excel表中的所有表
# sheet1 = workbook.active
# # print(sheet1['A1'].value)

workbook2 = openpyxl.load_workbook('temp.xlsx')	# 返回一个workbook数据类型的值
sheet = workbook2['Sheet1']
row=1
for i in sheet['H']:
    try:
        if len(i.value)>=3 and i.value[-3]=='(' and i.value[-1]==')':
            value=i.value[:-3]
            sheet['H'+str(row)].value=value
        elif len(i.value)>=4 and i.value[-4]=='(' and i.value[-1]==')':
            value=i.value[:-4]
            sheet['H'+str(row)].value=value
        elif len(i.value)>=6 and  i.value[-6]=='(' and i.value[-1]==')':
            value=i.value[:-6]
            sheet['H'+str(row)].value=value
        elif len(i.value)>=5 and  i.value[-5]=='(' and i.value[-1]==')':
            value=i.value[:-5]
            sheet['H'+str(row)].value=value
    # sheet['H1']=5
    except:
        break
    print(row)
    row+=1
workbook2.save('xx.xlsx')
print(1)